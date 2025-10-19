"""
Faraday Shield Analyser - Native Android/Desktop Kivy App
Entry point for both Android APK and Desktop
"""
import os
import json
from pathlib import Path
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserIconView
from kivy.core.window import Window
from kivy.utils import platform
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Set up data directory based on platform
if platform == 'android':
    from android.storage import app_storage_path
    DATA_DIR = Path(app_storage_path())
else:
    DATA_DIR = Path(__file__).parent

EXPERIMENTS_FILE = DATA_DIR / "experiments.json"
CREDS_FILE = DATA_DIR / "creds.json"

# Ensure data files exist
os.makedirs(DATA_DIR, exist_ok=True)
if not CREDS_FILE.exists():
    CREDS_FILE.write_text(json.dumps({"admin": "admin123"}))
if not EXPERIMENTS_FILE.exists():
    EXPERIMENTS_FILE.write_text(json.dumps([]))


class LoginScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'login'
        
        layout = BoxLayout(orientation='vertical', padding=50, spacing=20)
        
        # Title
        title = Label(
            text='Faraday Shield Analyser',
            size_hint=(1, 0.3),
            font_size='24sp',
            bold=True
        )
        
        # Username
        self.username_input = TextInput(
            hint_text='Username',
            multiline=False,
            size_hint=(1, 0.1)
        )
        
        # Password
        self.password_input = TextInput(
            hint_text='Password',
            password=True,
            multiline=False,
            size_hint=(1, 0.1)
        )
        
        # Login button
        login_btn = Button(
            text='Login',
            size_hint=(1, 0.15),
            background_color=(0.2, 0.6, 1, 1)
        )
        login_btn.bind(on_press=self.do_login)
        
        # Error label
        self.error_label = Label(
            text='',
            size_hint=(1, 0.1),
            color=(1, 0, 0, 1)
        )
        
        layout.add_widget(title)
        layout.add_widget(Label(size_hint=(1, 0.1)))  # Spacer
        layout.add_widget(self.username_input)
        layout.add_widget(self.password_input)
        layout.add_widget(login_btn)
        layout.add_widget(self.error_label)
        
        self.add_widget(layout)
    
    def do_login(self, instance):
        username = self.username_input.text
        password = self.password_input.text
        
        try:
            with open(CREDS_FILE) as f:
                creds = json.load(f)
            
            if username in creds and creds[username] == password:
                self.manager.current = 'main'
                self.username_input.text = ''
                self.password_input.text = ''
                self.error_label.text = ''
            else:
                self.error_label.text = 'Invalid credentials'
        except Exception as e:
            self.error_label.text = f'Error: {str(e)}'


class ExperimentListScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = 'main'
        
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Header
        header = BoxLayout(size_hint=(1, 0.1), spacing=10)
        title = Label(text='Experiments', font_size='20sp', bold=True)
        logout_btn = Button(
            text='Logout',
            size_hint=(0.3, 1),
            background_color=(0.8, 0.2, 0.2, 1)
        )
        logout_btn.bind(on_press=self.do_logout)
        header.add_widget(title)
        header.add_widget(logout_btn)
        
        # Buttons row
        btn_row = BoxLayout(size_hint=(1, 0.1), spacing=10)
        
        new_exp_btn = Button(
            text='New Experiment',
            background_color=(0.2, 0.8, 0.2, 1)
        )
        new_exp_btn.bind(on_press=self.new_experiment)
        
        import_btn = Button(
            text='Import Excel',
            background_color=(0.2, 0.6, 1, 1)
        )
        import_btn.bind(on_press=self.import_excel)
        
        btn_row.add_widget(new_exp_btn)
        btn_row.add_widget(import_btn)
        
        # Experiments list
        self.exp_list = GridLayout(
            cols=1,
            spacing=5,
            size_hint_y=None
        )
        self.exp_list.bind(minimum_height=self.exp_list.setter('height'))
        
        scroll = ScrollView(size_hint=(1, 0.8))
        scroll.add_widget(self.exp_list)
        
        layout.add_widget(header)
        layout.add_widget(btn_row)
        layout.add_widget(scroll)
        
        self.add_widget(layout)
    
    def on_enter(self):
        self.load_experiments()
    
    def load_experiments(self):
        self.exp_list.clear_widgets()
        
        try:
            with open(EXPERIMENTS_FILE) as f:
                experiments = json.load(f)
            
            if not experiments:
                self.exp_list.add_widget(
                    Label(text='No experiments yet. Create one!', size_hint_y=None, height=40)
                )
            else:
                for exp in experiments:
                    exp_btn = Button(
                        text=f"{exp['name']} ({len(exp.get('data', []))} rows)",
                        size_hint_y=None,
                        height=60,
                        background_color=(0.3, 0.3, 0.3, 1)
                    )
                    exp_btn.bind(on_press=lambda x, e=exp: self.open_experiment(e))
                    self.exp_list.add_widget(exp_btn)
        except Exception as e:
            self.exp_list.add_widget(
                Label(text=f'Error loading: {str(e)}', size_hint_y=None, height=40)
            )
    
    def do_logout(self, instance):
        self.manager.current = 'login'
    
    def new_experiment(self, instance):
        # Show popup to create new experiment
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        name_input = TextInput(hint_text='Experiment Name', multiline=False)
        locations_input = TextInput(hint_text='Number of Locations', input_filter='int', multiline=False)
        frequencies_input = TextInput(hint_text='Number of Frequencies', input_filter='int', multiline=False)
        
        btn_box = BoxLayout(size_hint=(1, 0.3), spacing=10)
        create_btn = Button(text='Create')
        cancel_btn = Button(text='Cancel')
        
        popup = Popup(
            title='New Experiment',
            content=content,
            size_hint=(0.9, 0.6)
        )
        
        def create(instance):
            try:
                name = name_input.text.strip()
                num_locations = int(locations_input.text or 1)
                num_frequencies = int(frequencies_input.text or 1)
                
                if not name:
                    return
                
                # Create experiment structure
                columns = ['Frequency', 'Reference']
                for i in range(num_locations):
                    columns.append(f'L{i+1}')
                    columns.append(f'L{i+1} - Shielding')
                
                data = []
                for i in range(num_frequencies):
                    row = {col: 0.0 for col in columns}
                    row['Frequency'] = 0.0
                    data.append(row)
                
                # Save experiment
                with open(EXPERIMENTS_FILE) as f:
                    experiments = json.load(f)
                
                experiments.append({
                    'name': name,
                    'columns': columns,
                    'data': data
                })
                
                with open(EXPERIMENTS_FILE, 'w') as f:
                    json.dump(experiments, f)
                
                popup.dismiss()
                self.load_experiments()
            except Exception as e:
                print(f"Error creating experiment: {e}")
        
        create_btn.bind(on_press=create)
        cancel_btn.bind(on_press=popup.dismiss)
        
        btn_box.add_widget(create_btn)
        btn_box.add_widget(cancel_btn)
        
        content.add_widget(name_input)
        content.add_widget(locations_input)
        content.add_widget(frequencies_input)
        content.add_widget(btn_box)
        
        popup.open()
    
    def import_excel(self, instance):
        # File chooser popup
        content = BoxLayout(orientation='vertical')
        filechooser = FileChooserIconView(filters=['*.xlsx', '*.xls'])
        
        btn_box = BoxLayout(size_hint=(1, 0.1), spacing=10)
        select_btn = Button(text='Import')
        cancel_btn = Button(text='Cancel')
        
        popup = Popup(
            title='Select Excel File',
            content=content,
            size_hint=(0.9, 0.9)
        )
        
        def do_import(instance):
            if filechooser.selection:
                try:
                    if not EXCEL_SUPPORT:
                        print("Excel support not available")
                        popup.dismiss()
                        return
                    
                    file_path = filechooser.selection[0]
                    
                    # Read Excel using openpyxl directly
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    # Get columns from first row
                    columns = [cell.value for cell in ws[1]]
                    
                    # Add shielding columns
                    new_columns = []
                    for col in columns:
                        new_columns.append(col)
                        if col not in ['Frequency', 'Reference'] and not str(col).endswith('- Shielding'):
                            new_columns.append(f'{col} - Shielding')
                    
                    # Read data rows
                    data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, col in enumerate(columns):
                            if i < len(row):
                                row_dict[col] = float(row[i]) if row[i] is not None else 0.0
                        
                        # Calculate shielding
                        for col in columns:
                            if col not in ['Frequency', 'Reference']:
                                shielding_col = f'{col} - Shielding'
                                row_dict[shielding_col] = row_dict.get('Reference', 0) - row_dict.get(col, 0)
                        
                        data.append(row_dict)
                    
                    # Save
                    exp_name = Path(file_path).stem
                    
                    with open(EXPERIMENTS_FILE) as f:
                        experiments = json.load(f)
                    
                    experiments.append({
                        'name': exp_name,
                        'columns': new_columns,
                        'data': data
                    })
                    
                    with open(EXPERIMENTS_FILE, 'w') as f:
                        json.dump(experiments, f)
                    
                    popup.dismiss()
                    self.load_experiments()
                except Exception as e:
                    print(f"Import error: {e}")
        
        select_btn.bind(on_press=do_import)
        cancel_btn.bind(on_press=popup.dismiss)
        
        btn_box.add_widget(select_btn)
        btn_box.add_widget(cancel_btn)
        
        content.add_widget(filechooser)
        content.add_widget(btn_box)
        
        popup.open()
    
    def open_experiment(self, experiment):
        # TODO: Navigate to experiment view
        print(f"Opening experiment: {experiment['name']}")


class FaradayShieldApp(App):
    def build(self):
        Window.clearcolor = (0.95, 0.95, 0.95, 1)
        
        sm = ScreenManager()
        sm.add_widget(LoginScreen())
        sm.add_widget(ExperimentListScreen())
        
        return sm


if __name__ == '__main__':
    FaradayShieldApp().run()
